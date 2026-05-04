# Streamlit Athlete Signup App (Dynamic dependent dropdowns)
# Generated on: 2026-03-01T11:27:08
#
# Key fix:
# - Removed st.form wrapper around dependent widgets so Event dropdown updates immediately
#   when Gender/Division changes.
#
# This app does NOT load the original Excel workbook.

import re
from datetime import date
from io import BytesIO

import openpyxl
import pandas as pd
import streamlit as st

APP_VARIANT = "public_entry_only"
import traceback as tb

import smtplib
from email.message import EmailMessage

import datetime as dt

from google_sheets_roster import load_roster, parse_dob, last4_from_nric
from google_sheets_writer import sync_entries_to_sheet
from google_sheets_reader import read_sheet_as_df


from reference_lists import (
    ENTRY_HEADERS,
    TEAM_CODES,
    get_team_name,
    COUNTRIES,
    get_events,  # gender+division -> [(event_name,event_code), ...] filtered (blacked-out excluded)
)

DIVISIONS = {1: 'Masters (30 & above)',
 2: 'Age (7-8)',
 3: 'Age (9-10)',
 4: 'Age (11-12)',
 5: 'Age (13-14)',
 6: 'Age (15-16)',
 7: 'Para',
 8: 'Age (5-6)',
 10: 'Age (17-18)'}

# Event options derived from 'SMTFA INT - SCHEDULE 6&7 JUN NT.xlsx', sheet 'NT Version'.
# Keys are (schedule_gender, division); values are (event_name, event_code/S-NO).
SCHEDULE_EVENT_OPTIONS = {('GIRLS', 8): [('60M', '83'), ('150M', '254')],
 ('BOYS', 8): [('60M', '84'), ('150M', '255')],
 ('GIRLS', 2): [('60M', '85'), ('150M', '256'), ('400M', '298')],
 ('BOYS', 2): [('60M', '86'), ('150M', '257'), ('400M', '299')],
 ('GIRLS', 3): [('80M', '81'),
                ('200M', '258'),
                ('400M', '300'),
                ('LONG JUMP (80 cm board)', '176'),
                ('POLE VAULT', '204')],
 ('BOYS', 3): [('80M', '82'),
               ('200M', '259'),
               ('400M', '301'),
               ('LONG JUMP (80 cm board)', '177'),
               ('POLE VAULT', '205')],
 ('GIRLS', 4): [('100M', '73'),
                ('200M', '260'),
                ('600M', '302'),
                ('POLE VAULT', '206'),
                ('LONG JUMP (50 cm board)', '222')],
 ('BOYS', 4): [('100M', '74'),
               ('200M', '261'),
               ('600M', '303'),
               ('POLE VAULT', '207'),
               ('LONG JUMP (50 cm board)', '285')],
 ('GIRLS', 5): [('100M', '75'), ('200M', '262'), ('800M', '252'), ('LONG JUMP', '296')],
 ('BOYS', 5): [('100M', '76'), ('200M', '263'), ('800M', '253'), ('LONG JUMP', '297')],
 ('GIRLS', 6): [('100M', '77'), ('200M', '264'), ('800M', '250'), ('LONG JUMP', '304')],
 ('BOYS', 6): [('100M', '78'), ('200M', '265'), ('800M', '251'), ('LONG JUMP', '305')],
 ('GIRLS', 10): [('100M', '79'), ('200M', '266'), ('800M', '248')],
 ('BOYS', 10): [('100M', '80'), ('200M', '267'), ('800M', '249')],
 ('MEN', 1): [('HAMMER', '97'),
              ('POLE VAULT', '209'),
              ('TRIPLE JUMP', '200'),
              ('SHOT PUT (7.26 KG)', '15'),
              ('5000M', '23'),
              ('100M', '39'),
              ('JAVELIN (800G)', '67'),
              ('110M HURDLES (99.1 CM)', '113'),
              ('400M', '119'),
              ('1500M', '141'),
              ('LONG JUMP', '158'),
              ('400H(91.4 CM) x 10', '210'),
              ('DISCUS (2 KG)', '223'),
              ('800M', '240'),
              ('200M', '277'),
              ('5000M WALK', '308'),
              ('HIGH JUMP', '87'),
              ('100M HURDLES (91.4 CM)', '107'),
              ('SHOT PUT (6 KG)', '132'),
              ('DISCUS (1.5KG)', '187'),
              ('400H(84 CM) x 10', '214'),
              ('JAVELIN (700G)', '227'),
              ('100M HURDLES (84 CM)', '105'),
              ('SHOT PUT (5 KG)', '129'),
              ('DISCUS (1 KG)', '189'),
              ('JAVELIN (600G)', '229'),
              ('300H (76.2 CM) x 7', '218'),
              ('SHOT PUT (4 KG)', '131'),
              ('3000M WALK', '184'),
              ('JAVELIN (500G)', '231'),
              ('80M HURDLES (76.2 CM)', '103'),
              ('200H (68.6 CM) x 5', '221'),
              ('80M HURDLES (68.6 CM)', '100'),
              ('4X100M RELAY', '307'),
              ('4X400M RELAY', '317')],
 ('WOMEN', 1): [('4X100M RELAY', '306'),
                ('4X400M RELAY', '316'),
                ('HAMMER', '98'),
                ('POLE VAULT', '208'),
                ('800M', '233'),
                ('5000M', '1'),
                ('JAVELIN (600G)', '35'),
                ('100M', '50'),
                ('HIGH JUMP', '60'),
                ('TRIPLE JUMP', '91'),
                ('400M', '134'),
                ('1500M', '150'),
                ('SHOT PUT(4 KG)', '162'),
                ('LONG JUMP', '169'),
                ('3000M WALK', '178'),
                ('DISCUS (1 KG)', '193'),
                ('200M', '270'),
                ('100M HURDLES (84 CM)', '104'),
                ('80M HURDLES (76.2 CM)', '101'),
                ('400H(76.2 CM) x 10', '216'),
                ('JAVELIN (500G)', '19'),
                ('SHOT PUT (3 KG)', '166'),
                ('300H (76.2 CM) x 7', '217'),
                ('80M HURDLES (68.6 CM)', '99'),
                ('200H (68.6 CM) x 5', '220')],
 ('PARA WOMEN', 7): [('100M', '71'), ('200M', '268'), ('LONG JUMP', '294')],
 ('PARA MEN', 7): [('100M', '72'), ('200M', '269'), ('LONG JUMP', '295')]}

IC_LAST4_RE = re.compile(r"^\d{3}[A-Za-z]$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def normalize_ic_last4(s: str) -> str:
    return (s or "").strip().upper()

def is_valid_ic_last4(s: str) -> bool:
    return bool(IC_LAST4_RE.match((s or "").strip()))

def normalize_email(s: str) -> str:
    return (s or "").strip()

def is_valid_email(s: str) -> bool:
    s = normalize_email(s)
    if not s or len(s) > 254:
        return False
    return bool(EMAIL_RE.match(s))


def _match_option_case_insensitive(value: str, options: list[str]) -> str:
    """Return the option whose casefold matches value; else empty string."""
    v = (value or "").strip()
    if not v:
        return ""
    vcf = v.casefold()
    for o in options:
        if (o or "").strip().casefold() == vcf:
            return o
    return ""

def compute_unique_id(first_name: str, ic_last4: str, dob: date) -> str:
    if not first_name or not ic_last4 or not dob:
        return ""
    ic = normalize_ic_last4(ic_last4)
    return f"{first_name.strip()[:1]}{ic[:4]}{dob.year % 100:02d}".upper()



def _event_sort_key(event_name: str):
    """Sort timed/running events by distance, then field events alphabetically."""
    name = str(event_name or "").strip()
    upper = name.upper()

    relay = re.search(r"\b(\d+)\s*[Xx]\s*(\d+(?:\.\d+)?)\s*M\b", upper)
    if relay:
        legs = float(relay.group(1))
        leg_distance = float(relay.group(2))
        return (0, legs * leg_distance, upper)

    dist = re.search(r"\b(\d+(?:\.\d+)?)\s*(KM|M)\b", upper)
    if dist:
        value = float(dist.group(1))
        metres = value * 1000 if dist.group(2) == "KM" else value
        return (0, metres, upper)

    return (1, float("inf"), upper)

def _schedule_gender_for_division(gender: str, division_no: int) -> str:
    """Map form gender to the schedule gender labels for the selected division.

    Accepts form values (Male/Female), code values (M/F), and schedule labels
    (MEN/WOMEN/BOYS/GIRLS/PARA MEN/PARA WOMEN).
    """
    d = int(division_no)
    raw = (gender or "").strip().upper()
    code = gender_to_code(gender)

    # If a schedule label was passed in directly, keep it when appropriate.
    if raw in ("MEN", "BOYS"):
        return "MEN" if d == 1 else ("PARA MEN" if d == 7 else "BOYS")
    if raw in ("WOMEN", "GIRLS"):
        return "WOMEN" if d == 1 else ("PARA WOMEN" if d == 7 else "GIRLS")
    if raw in ("PARA MEN", "PARA WOMEN"):
        return raw

    if code == "M":
        if d == 1:
            return "MEN"
        if d == 7:
            return "PARA MEN"
        return "BOYS"

    if code == "F":
        if d == 1:
            return "WOMEN"
        if d == 7:
            return "PARA WOMEN"
        return "GIRLS"

    return ""


def allowed_events(gender: str, division_no: int):
    """Return list of (event_name, event_code) from the SMTFA schedule.

    If gender is blank/unrecognised, return the union of events for that division so
    the dropdown is not empty while the user is still completing the form.
    """
    d = int(division_no)
    schedule_gender = _schedule_gender_for_division(gender, d)

    exact = list(SCHEDULE_EVENT_OPTIONS.get((schedule_gender, d), []))
    if exact:
        return exact

    # Fallback: show all unique event names for the selected division.
    # This prevents the Select Event dropdown from being empty when gender is blank
    # or when a schedule label is unexpected. The current gender still determines
    # the correct event code once selected.
    merged = []
    seen = set()
    for (_gender_key, _division_key), _opts in SCHEDULE_EVENT_OPTIONS.items():
        if int(_division_key) != d:
            continue
        for _name, _code in _opts:
            if _name not in seen:
                merged.append((_name, _code))
                seen.add(_name)
    return merged

def export_entries_to_excel(header_info: dict, entries: pd.DataFrame) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entry Form"

    ws["A1"].value = "Team Name"
    ws["B1"].value = header_info.get("team_name", "")

    ws["A2"].value = "Billing Contact Name"
    ws["B2"].value = header_info.get("billing_name", "")

    ws["A3"].value = "Billing Email"
    ws["B3"].value = header_info.get("billing_email", "")

    ws["A4"].value = "Charge Code"
    ws["B4"].value = header_info.get("charge_code", "")

    ws["A5"].value = "P/O to be sent"
    ws["B5"].value = header_info.get("po_to_be_sent", "")

    header_row = 7
    for c, h in enumerate(ENTRY_HEADERS, start=1):
        ws.cell(header_row, c).value = h

    start_row = header_row + 1
    for idx, row in entries.reset_index(drop=True).iterrows():
        r = start_row + idx
        ws.cell(r, 1).value = idx + 1  # No
        ws.cell(r, 2).value = row.get("last_name", "")
        ws.cell(r, 3).value = row.get("first_name", "")
        ws.cell(r, 4).value = row.get("gender", "")
        ws.cell(r, 5).value = row.get("birth_date", None)
        ws.cell(r, 6).value = row.get("ic_last4", "")
        ws.cell(r, 7).value = row.get("unique_id", "")
        ws.cell(r, 8).value = row.get("nationality", "")
        ws.cell(r, 9).value = row.get("contact_number", "")
        ws.cell(r,10).value = row.get("email", "")
        ws.cell(r,11).value = row.get("team_code", "")
        ws.cell(r,12).value = row.get("team_name", "")
        ws.cell(r,13).value = row.get("event_code", "")
        ws.cell(r,14).value = row.get("event_division", "")
        ws.cell(r,15).value = row.get("season_best", "")
        ws.cell(r,16).value = row.get("emergency_contact_name", "")
        ws.cell(r,17).value = row.get("emergency_contact_number", "")
        ws.cell(r,18).value = row.get("coach_full_name", "")
        ws.cell(r,19).value = row.get("parq", "")

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
# ---------------- UI ----------------
st.set_page_config(page_title="SMTFA International Masters T&F Signup", layout="wide")
st.title("SMTFA International Masters T&F Signup")

def _apply_pending_text_updates():
    """Apply any pending text updates BEFORE widgets are instantiated."""
    pending = [k for k in list(st.session_state.keys()) if k.endswith("__pending")]
    for pk in pending:
        base = pk[:-9]  # strip '__pending'
        st.session_state[base] = st.session_state.get(pk, "")
        try:
            del st.session_state[pk]
        except Exception:
            pass

_apply_pending_text_updates()


# ---------------- Hidden configuration (no sidebar controls) ----------------
ROSTER_SHEET_URL = st.secrets.get(
    "ROSTER_SHEET_URL",
    "https://docs.google.com/spreadsheets/d/1PnTKatJGW3Eazy6YpDqnRHVZrVLRvK9rA_vBIVXKfUU/edit?usp=sharing",
)
ROSTER_WORKSHEET = st.secrets.get("ROSTER_WORKSHEET", "")

OUTPUT_SHEET_URL = st.secrets.get(
    "OUTPUT_SHEET_URL",
    "https://docs.google.com/spreadsheets/d/11AxxxJkO5CGqCjMg6gyAwW1BlegJ0Wf6G7F60rN8X0g/edit?usp=sharing",
)
OUTPUT_WORKSHEET = st.secrets.get("OUTPUT_WORKSHEET", "")

st.session_state.setdefault("roster_sheet_url", ROSTER_SHEET_URL)
st.session_state.setdefault("roster_worksheet", ROSTER_WORKSHEET)
st.session_state["use_roster"] = True

st.session_state.setdefault("output_sheet_url", OUTPUT_SHEET_URL)
st.session_state.setdefault("output_worksheet", OUTPUT_WORKSHEET)
st.session_state["sync_enabled"] = True

# Preload roster once per session
if st.session_state.get("use_roster") and (st.session_state.get("roster_sheet_url") or "").strip():
    if "roster_cache_rows" not in st.session_state:
        try:
            st.session_state["roster_cache_rows"] = load_roster(
                st.session_state.get("roster_sheet_url", ""),
                worksheet=((st.session_state.get("roster_worksheet") or "").strip() or None),
            )
        except Exception as e:
            st.session_state["roster_cache_rows"] = []
            st.session_state["roster_cache_error"] = f"{type(e).__name__}: {repr(e)}"

if st.session_state.get("roster_cache_error"):
    st.warning(f"Roster could not be loaded. Name matching may be unavailable. ({st.session_state['roster_cache_error']})")
# ---------------------------------------------------------------------------


def _normalize_header(h: str) -> str:
    h = (h or "").strip().casefold()
    h = re.sub(r"[^a-z0-9]+", "_", h).strip("_")
    return h


def gender_to_code(g: str) -> str:
    g = (g or "").strip().casefold()
    if g in ("m", "male"):
        return "M"
    if g in ("f", "female"):
        return "F"
    return ""

def code_to_gender_display(code: str) -> str:
    c = (code or "").strip().upper()
    if c == "M":
        return "Male"
    if c == "F":
        return "Female"
    return ""

def safe_date_max(v):
    """Return a max_value for st.date_input that is always >= v (prevents StreamlitAPIException)."""
    mx = dt.date.today()
    try:
        if isinstance(v, dt.date) and v > mx:
            return v
    except Exception:
        pass
    return mx


def send_confirmation_email_smtp(to_email: str, subject: str, body: str) -> None:
    """Send a confirmation email via SMTP using Streamlit secrets.

    Required secrets:
      SMTP_HOST, SMTP_PORT (optional, default 587), SMTP_USER, SMTP_PASS
    Optional secrets:
      SMTP_FROM (display name + email) e.g. "SAA Entries <your@email>"
    """
    host = (st.secrets.get("SMTP_HOST", "") or "").strip()
    user = (st.secrets.get("SMTP_USER", "") or "").strip()
    password = (st.secrets.get("SMTP_PASS", "") or "").strip()
    sender = (st.secrets.get("SMTP_FROM", "") or user).strip()
    port = int(st.secrets.get("SMTP_PORT", 587) or 587)

    if not host or not user or not password:
        raise RuntimeError("SMTP secrets are missing (SMTP_HOST/SMTP_USER/SMTP_PASS).")

    msg = EmailMessage()
    msg["From"] = sender
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body)

    with smtplib.SMTP(host, port, timeout=20) as smtp:
        smtp.ehlo()
        smtp.starttls()
        smtp.ehlo()
        smtp.login(user, password)
        smtp.send_message(msg)


def build_semicolon_export_from_output_sheet(sheet_df: pd.DataFrame, record_type: str = "I") -> str:
    """Build semicolon-delimited export rows (no header) from the OUTPUT sheet.

    I format keeps the existing fixed-width layout:
      I; last_name; first_name; ; gender; dob; team_code; team_name; ; ... nationality ... unique_id ...

    E format follows the uploaded sample:
      E; last_name; first_name; ; gender; dob; team_code; team_name; ; ; event_code; season_best; ; M; division; ; ;
    """
    if sheet_df is None or sheet_df.empty:
        return ""

    col_map = {_normalize_header(c): c for c in sheet_df.columns}

    def get(row, key, default=""):
        c = col_map.get(key)
        return row.get(c, default) if c else default

    def fmt_date(v):
        if v is None:
            return ""
        try:
            d = parse_dob(v)
            return d.strftime("%d/%m/%Y") if d else ""
        except Exception:
            return str(v).strip()

    def fmt_gender(v):
        raw = str(v or "").strip()
        return gender_to_code(raw) or raw

    def fmt_division(v):
        s = str(v or "").strip()
        if not s:
            return ""
        try:
            f = float(s)
            if f.is_integer():
                return str(int(f))
        except Exception:
            pass
        return s

    def export_event_code(event_name, stored_event_code):
        """Return the E-file event code shown in the uploaded sample.

        The schedule S/NO may be stored internally as event_code, but the E file sample
        expects codes such as 100, 200, 800, 100H, HJ, LJ, DT.
        """
        name = str(event_name or "").strip().upper()
        stored = str(stored_event_code or "").strip().upper()

        # If stored code already looks like an event code (not merely a schedule S/NO),
        # keep it. Examples: 100H, HJ, LJ, DT.
        if stored and not stored.isdigit():
            return stored

        # Running / hurdle events
        m = re.search(r"(\d+)\s*M", name)
        if m:
            distance = m.group(1)
            if "HURD" in name:
                return f"{distance}H"
            if "WALK" in name:
                return f"{distance}W"
            return distance

        # Field-event codes commonly used in Meet Manager imports
        if "HIGH JUMP" in name:
            return "HJ"
        if "LONG JUMP" in name:
            return "LJ"
        if "TRIPLE JUMP" in name:
            return "TJ"
        if "POLE VAULT" in name:
            return "PV"
        if "SHOT" in name:
            return "SP"
        if "DISCUS" in name:
            return "DT"
        if "JAVELIN" in name:
            return "JT"
        if "HAMMER" in name:
            return "HT"

        relay = re.search(r"(\d+)\s*[Xx]\s*(\d+)", name)
        if relay:
            return f"{relay.group(1)}X{relay.group(2)}"

        return stored or name

    rt = (record_type or "I").strip().upper()[:1] or "I"
    if rt not in ("I", "E"):
        rt = "I"

    lines = []
    for _, row in sheet_df.iterrows():
        last_name = str(get(row, "last_name", "")).strip()
        first_name = str(get(row, "first_name", "")).strip()
        gender = fmt_gender(get(row, "gender", ""))
        dob = fmt_date(get(row, "dob", "") or get(row, "birth_date", "") or get(row, "date_of_birth", ""))
        team_code = str(get(row, "team_code", "")).strip()
        team_name = str(get(row, "team_name", "")).strip()
        nationality = str(get(row, "nationality", "")).strip()
        unique_id = str(get(row, "unique_id", "")).strip()
        event_name = str(get(row, "event", "")).strip()
        event_code_stored = str(get(row, "event_code", "")).strip()
        event_code_export = export_event_code(event_name, event_code_stored)
        season_best = str(get(row, "season_best", "")).strip()
        division = fmt_division(get(row, "event_division", "") or get(row, "division", ""))

        if rt == "E":
            fields = [
                "E",
                last_name,
                first_name,
                "",
                gender,
                dob,
                team_code,
                team_name,
                "",
                "",
                event_code_export,
                season_best,
                "",
                "M",
                division,
                "",
                "",
            ]
            include_row = any([last_name, first_name, gender, dob, team_code, team_name, event_code_export, season_best, division])
        else:
            fields = [
                "I",
                last_name,
                first_name,
                "",
                gender,
                dob,
                team_code,
                team_name,
                "", "", "", "", "", "", "", "",
                nationality,
                "", "", "", "",
                unique_id,
                "", "", "",
            ]
            include_row = any([last_name, first_name, gender, dob, team_code, team_name, nationality, unique_id])

        if not include_row:
            continue

        lines.append("; ".join(fields))

    return "\n".join(lines) + ("\n" if lines else "")


def _sheet_df_to_entries(df: pd.DataFrame) -> list[dict]:
    if df is None or df.empty:
        return []
    col_map = {_normalize_header(c): c for c in df.columns}

    def get(row, key, default=""):
        c = col_map.get(key)
        return row.get(c, default) if c else default

    entries = []
    for _, row in df.iterrows():
        first_name = str(get(row, "first_name", "") or get(row, "firstname", "") or get(row, "first", "")).strip()
        other_name = str(get(row, "other_name", "") or get(row, "othername", "")).strip()
        last_name = str(get(row, "last_name", "") or get(row, "lastname", "") or get(row, "last", "")).strip()
        gender_raw = str(get(row, "gender", "")).strip()
        gcode = gender_to_code(gender_raw)
        gender = code_to_gender_display(gcode) or gender_raw

        dob_raw = get(row, "birth_date", "") or get(row, "dob", "") or get(row, "date_of_birth", "")
        try:
            birth_date = parse_dob(dob_raw)
        except Exception:
            birth_date = None

        nric_raw = str(get(row, "nric", "") or "").strip()
        ic_last4 = str(get(row, "ic_last4", "") or "").strip()
        if not ic_last4 and nric_raw:
            ic_last4 = last4_from_nric(nric_raw)
        ic_last4 = normalize_ic_last4(ic_last4)

        nationality = str(get(row, "nationality", "")).strip()
        singapore_pr = str(
            get(row, "singapore_pr", "")
            or get(row, "sg_pr", "")
            or get(row, "singapore_pr_status", "")
            or get(row, "singapore_pr?", "")
            or ""
        ).strip()
        if singapore_pr.casefold() in ("true", "1", "y", "yes"):
            singapore_pr = "Yes"
        elif singapore_pr.casefold() in ("false", "0", "n", "no"):
            singapore_pr = "No"
        else:
            singapore_pr = singapore_pr or "No"
        unique_id = str(get(row, "unique_id", "")).strip()
        team_name = str(get(row, "team_name", "")).strip()
        team_code = str(get(row, "team_code", "")).strip()
        event = str(get(row, "event", "")).strip()
        event_code = str(get(row, "event_code", "")).strip()
        season_best = str(get(row, "season_best", "") or get(row, "season best", "") or "").strip()
        parq = str(get(row, "parq", "") or get(row, "par_q", "") or get(row, "par-q", "") or "").strip()
        emergency_contact_name = str(get(row, "emergency_contact_name", "") or get(row, "emergency contact name", "") or "").strip()
        emergency_contact_number = str(get(row, "emergency_contact_number", "") or get(row, "emergency contact number", "") or "").strip()
        coach_full_name = str(get(row, "coach_full_name", "") or get(row, "coach full name", "") or "").strip()
        charge_code = str(get(row, "charge_code", "")).strip()
        po_to_be_sent = str(get(row, "po_to_be_sent", "")).strip()
        email = normalize_email(str(get(row, "email", "")).strip())
        contact_number = str(get(row, "contact_number", "") or get(row, "contact", "")).strip()

        full_name_sheet = str(get(row, "full_name", "") or get(row, "full", "") or "").strip()
        name = " ".join([p for p in [first_name, other_name, last_name] if p]).strip()
        if full_name_sheet:
            name = name or full_name_sheet

        if not any([name, unique_id, team_code, team_name, email, contact_number, ic_last4]):
            continue

        entries.append({
            "name": name,
            "full_name": (full_name_sheet or name),
            "name_passport": str(get(row, "name_passport", "") or get(row, "name_as_per_nric_passport", "") or "").strip() or (full_name_sheet or name),
            "first_name": first_name,
            "other_name": other_name,
            "last_name": last_name,
            "gender": gender if gender in ("Male","Female","M","F") else "",
            "birth_date": birth_date,
            "ic_last4": ic_last4,
            "nationality": nationality,
            "singapore_pr": singapore_pr,
            "unique_id": unique_id,
            "team_name": team_name,
            "team_code": team_code,
            "event": event,
            "event_code": event_code,
            "charge_code": charge_code,
            "po_to_be_sent": po_to_be_sent if po_to_be_sent in ("Yes","No") else "",
            "season_best": season_best,
            "parq": parq if parq in ("Y", "N") else "Y",
            "emergency_contact_name": emergency_contact_name,
            "emergency_contact_number": emergency_contact_number,
            "coach_full_name": coach_full_name,
            "email": email,
            "contact_number": contact_number,
        })
    return entries

# Preload existing entries from OUTPUT Google Sheet on app start (only if empty)
st.session_state.setdefault("entries", [])
st.session_state.setdefault("full_name", "")
st.session_state.setdefault("name_passport", "")

if not st.session_state.entries:
    _preload_url = (st.session_state.get("output_sheet_url") or "").strip()
    _preload_ws = (st.session_state.get("output_worksheet") or "").strip() or None
    if _preload_url:
        try:
            _df_pre = read_sheet_as_df(_preload_url, worksheet=_preload_ws)
            st.session_state.entries = _sheet_df_to_entries(_df_pre)
        except Exception as e:
            st.session_state["preload_error"] = f"{type(e).__name__}: {repr(e)}"

if st.session_state.get("preload_error"):
    st.warning(f"Could not preload existing entries from output sheet. ({st.session_state['preload_error']})")



if "entries" not in st.session_state:
    st.session_state.entries = []

with st.sidebar:
    st.header("Team / Billing")
    default_team_code = st.selectbox("Default Team Code (optional)", [""] + TEAM_CODES, index=0, key="default_team_code")
    default_team_name = get_team_name(default_team_code) if default_team_code else ""
    team_name_header = st.text_input("Default Team Name (for header)", value=default_team_name, key="team_name_header")
    billing_name = st.text_input("Billing contact name", value="", key="billing_name")
    billing_email = st.text_input("Billing email", value="", key="billing_email")
    charge_code = st.text_input("Charge code (optional)", value="", key="charge_code")
    po_to_be_sent = st.radio("P/O to be sent", options=["No", "Yes"], index=0, horizontal=True, key="po_to_be_sent")
    if billing_email and not is_valid_email(billing_email):
        st.warning("Billing email looks invalid. Please double-check it.")

# Defensive: ensure billing fields are bound even if sidebar UI is modified
po_to_be_sent = st.session_state.get("po_to_be_sent", "No")
charge_code = st.session_state.get("charge_code", "")
st.subheader("Athlete Entry Form")

# Athlete fields (no form, so dependent dropdowns update immediately)
c1, c2, c3, c4 = st.columns(4)
with c1:
    st.text_input("Last Name", key="last_name")
with c2:
    st.text_input("First Name", key="first_name")
with c3:
    st.text_input("Other Name (optional)", key="other_name")
with c4:
    gender = st.selectbox("Gender", ["", "Male", "Female"], index=0, key="gender")

# Name as per NRIC/Passport is a separate required field.
# It is intentionally NOT auto-filled from First Name / Last Name / Full Name.
passport_name = st.text_input("Name as per NRIC/Passport", key="name_passport")
passport_ok = bool((passport_name or "").strip())
if not passport_ok:
    st.warning("Name as per NRIC/Passport is required.")

# Live validation: gender (mandatory)
gender_ok = gender in ('Male','Female')
if not gender_ok:
    st.warning("Gender is required (select Male or Female).")


last_name = st.session_state.get("last_name", "")
first_name = st.session_state.get("first_name", "")
other_name = st.session_state.get("other_name", "")

# If user edits any name fields after selecting from roster, clear roster-derived FULL_NAME and UNIQUE_ID
current_name_sig = "|".join([
    (st.session_state.get("first_name", "") or "").strip(),
    (st.session_state.get("other_name", "") or "").strip(),
    (st.session_state.get("last_name", "") or "").strip(),
])
prev_sig = (st.session_state.get("full_name_signature", "") or "").strip()
if prev_sig and current_name_sig != prev_sig:
    # User typed a new name; clear roster-derived fields so they don't persist
    st.session_state["full_name__pending"] = ""
    st.session_state["unique_id_override__pending"] = ""
    st.session_state["db_name_override__pending"] = ""
    st.session_state["full_name_signature__pending"] = ""
    st.rerun()


# Roster match selector (Google Sheet) — selecting a row fills fields (no splitting)
search_text = (" ".join([p for p in [first_name, other_name, last_name] if (p or "").strip()])).strip()

_roster_enabled = bool(st.session_state.get("use_roster"))
_roster_url = (st.session_state.get("roster_sheet_url") or "").strip()

# Small inline hints so you can see why the dropdown may not appear
if not _roster_enabled:
    st.caption("Roster search is OFF (check configuration).")
elif not _roster_url:
    st.caption("Roster sheet URL is empty (set ROSTER_SHEET_URL in secrets).")
elif len(search_text) < 2:
    st.caption("Type at least 2 characters in First/Other/Last name to search the roster.")

matches = []
roster_rows = []
if _roster_enabled and _roster_url and len(search_text) >= 2:
    try:
        roster_rows = st.session_state.get("roster_cache_rows")
        # If cache exists but is empty, try one reload per session (handles first-run load glitches)
        if isinstance(roster_rows, list) and (len(roster_rows) == 0) and (not st.session_state.get("roster_cache_reloaded_once")):
            st.session_state["roster_cache_reloaded_once"] = True
            roster_rows = load_roster(
                st.session_state.get("roster_sheet_url", ""),
                worksheet=((st.session_state.get("roster_worksheet") or "").strip() or None),
            )
            st.session_state["roster_cache_rows"] = roster_rows
        if not isinstance(roster_rows, list):
            roster_rows = load_roster(
                st.session_state.get("roster_sheet_url", ""),
                worksheet=((st.session_state.get("roster_worksheet") or "").strip() or None),
            )
            st.session_state["roster_cache_rows"] = roster_rows
    except Exception as e:
        st.error(f"Roster load error: {type(e).__name__}: {repr(e)}")
        roster_rows = []

    st.caption(f"Roster loaded: {len(roster_rows)} rows")
    q = search_text.casefold()
    for r in roster_rows:
        full_name = str(r.get("FULL_NAME", "") or "")
        fn = str(r.get("FIRST_NAME", "") or "")
        ln = str(r.get("LAST_NAME", "") or "")
        on = str(r.get("OTHER_NAME", "") or "")
        team = str(r.get("TEAM_NAME", "") or "")
        uid = str(r.get("UNIQUE_ID", "") or "")
        nric = str(r.get("NRIC", "") or "")
        # Match on tokens across ALL name fields (FIRST/LAST/OTHER/FULL), plus team/uid/nric(last4)
        full_name = str(r.get("FULL_NAME", "") or "")
        name_hay = " ".join([full_name, fn, on, ln]).casefold()
        tokens = [t.casefold() for t in search_text.split() if t.strip()]
        extra_hay = " ".join([team, uid, last4_from_nric(nric)]).casefold()
        # Scored OR-matching: show suggestions even if only part of the name is typed
        score = 0
        if tokens:
            score += sum(1 for t in tokens if t in name_hay)
        if q and q in name_hay:
            score += 2  # boost full-query name hits
        if q and q in extra_hay:
            score += 1
        if score > 0:
            matches.append((score, r))

    if matches:
        matches = [r for _, r in sorted(matches, key=lambda x: x[0], reverse=True)]
    st.caption(f"Matches found: {len(matches)}")

    if roster_rows and not matches:
        st.info(f"No roster matches for: '{search_text}'. You can refine the search (try first name, last name, team, UID, or NRIC last-4).")

    # Optional browse mode (helps confirm data is loading)
    browse_mode = st.toggle("Browse roster (show first 25)", value=False, key="browse_roster_mode")
    if browse_mode and roster_rows:
        matches = roster_rows[:25]

    if matches:
        labels = []
        for r in matches[:25]:
            fn = str(r.get("FIRST_NAME", "") or "").strip()
            fn_raw = str(r.get("FIRST_NAME", "") or "").strip()
            ln = str(r.get("LAST_NAME", "") or "").strip()
            on = str(r.get("OTHER_NAME", "") or "").strip()
            nric = str(r.get("NRIC", "") or "").strip()
            dob_val = parse_dob(r.get("DOB"))
            # Privacy: only show birth year in roster match/browse labels, not full DOB.
            if hasattr(dob_val, "strftime") and dob_val:
                dob_str = dob_val.strftime("%Y")
            else:
                _dob_raw = str(r.get("DOB", "") or "").strip()
                _year_match = re.search(r"(?:19|20)\d{2}", _dob_raw)
                dob_str = _year_match.group(0) if _year_match else ""
            gen = str(r.get("GENDER", "") or "").strip()
            nat = str(r.get("NATIONALITY", "") or "").strip()
            uid = str(r.get("UNIQUE_ID", "") or "").strip()
            tcode = str(r.get("TEAM_CODE", "") or "").strip()
            team = str(r.get("TEAM_NAME", "") or "").strip()

            # Privacy: roster dropdown labels show first name only and redact last name.
            # Underlying roster row still contains full details for autofill after selection.
            first_for_label = fn or str(r.get("FIRST_NAME", "") or "").strip()
            last_for_label = ln or str(r.get("LAST_NAME", "") or "").strip()
            label_parts = []
            if first_for_label:
                label_parts.append(first_for_label)
            if last_for_label:
                label_parts.append("*")
            label = " ".join(label_parts).strip() or "(unnamed roster entry)"
            parts = []
            n4 = last4_from_nric(nric)
            if n4:
                parts.append(f"NRIC(last4): {n4}")
            if dob_str:
                parts.append(f"{dob_str}")
            if gen:
                parts.append(f"{gen}")
            if nat:
                parts.append(f"{nat}")
            team_piece = " ".join([p for p in [tcode, team] if p]).strip()
            if team_piece:
                parts.append(f"{team_piece}")
            if parts:
                label = label + " | " + " | ".join(parts)
            labels.append(label)

        sel_key = "athlete_roster_match"
        options = ["(keep typed)"] + list(range(len(labels)))
        chosen = st.selectbox(
            "Select From List of Matches :",
            options=options,
            key=sel_key,
            format_func=lambda x: "(keep typed)" if x == "(keep typed)" else labels[int(x)],
        )

        if chosen != "(keep typed)":
            idx = int(chosen)
            r = matches[idx]

            full_name_sel = str(r.get("FULL_NAME", "") or "").strip()
            if not full_name_sel:
                fn_tmp = str(r.get("FIRST_NAME", "") or "").strip()
                on_tmp = str(r.get("OTHER_NAME", "") or "").strip()
                ln_tmp = str(r.get("LAST_NAME", "") or "").strip()
                full_name_sel = " ".join([p for p in [fn_tmp, on_tmp, ln_tmp] if p]).strip()

            fn = str(r.get("FIRST_NAME", "") or "").strip()
            fn_raw = str(r.get("FIRST_NAME", "") or "").strip()
            ln = str(r.get("LAST_NAME", "") or "").strip()
            on = str(r.get("OTHER_NAME", "") or "").strip()
            nric = str(r.get("NRIC", "") or "").strip()
            dob = parse_dob(r.get("DOB"))
            gender_raw = str(r.get("GENDER", "") or "").strip().upper()
            nat_raw = str(r.get("NATIONALITY", "") or "").strip()
            sgpr_raw = str(r.get("SINGAPORE_PR", "") or r.get("SG_PR", "") or r.get("PR_STATUS", "") or "").strip()
            uid = str(r.get("UNIQUE_ID", "") or "").strip()
            tname = str(r.get("TEAM_NAME", "") or "").strip()
            tcode_raw = str(r.get("TEAM_CODE", "") or "").strip()

            # Populate name fields
            st.session_state["first_name__pending"] = fn_raw or on
            st.session_state["last_name__pending"] = ln
            st.session_state["other_name__pending"] = on
            st.session_state["full_name__pending"] = full_name_sel
            roster_name_passport = str(
                r.get("NAME_PASSPORT", "")
                or r.get("NAME_AS_PER_NRIC_PASSPORT", "")
                or r.get("NAME AS PER NRIC/PASSPORT", "")
                or ""
            ).strip()
            if roster_name_passport:
                st.session_state["name_passport__pending"] = roster_name_passport
            st.session_state["full_name_signature__pending"] = "|".join([
                (st.session_state.get("first_name__pending", "") or "").strip(),
                (st.session_state.get("other_name__pending", "") or "").strip(),
                (st.session_state.get("last_name__pending", "") or "").strip(),
            ])

            # Populate other fields
            st.session_state["ic_last4__pending"] = last4_from_nric(nric)
            st.session_state["birth_date__pending"] = dob
            # Gender from roster (may be blank; still mandatory to submit)
            if gender_raw in ("M","F","MALE","FEMALE"):
                st.session_state["gender__pending"] = ("Male" if gender_raw.startswith("M") else "Female")
            else:
                st.session_state["gender__pending"] = ""

            # Nationality: if not in list, store as override so it still appears in the dropdown
            nat_pick = _match_option_case_insensitive(nat_raw, (COUNTRIES or []))
            if nat_pick:
                st.session_state["nationality__pending"] = nat_pick
                st.session_state["nationality_override__pending"] = ""
            else:
                st.session_state["nationality__pending"] = nat_raw
                st.session_state["nationality_override__pending"] = nat_raw

            _nat_cf = nat_raw.casefold()
            _sgpr_cf = sgpr_raw.casefold()
            roster_is_sg_pr = (
                _sgpr_cf in ("yes", "y", "true", "1", "pr", "singapore pr", "sg pr")
                or _nat_cf in ("singapore pr", "sg pr")
            )
            st.session_state["singapore_pr__pending"] = bool(roster_is_sg_pr)
            if roster_is_sg_pr and not nat_pick:
                # Keep nationality as an IOC/WA code while using the checkbox for PR status.
                st.session_state["nationality__pending"] = "SGP"
                st.session_state["nationality_override__pending"] = ""

            # Unique ID override from roster
            st.session_state["unique_id_override__pending"] = uid
            # Optional roster fields, if available
            roster_email = str(r.get("EMAIL", "") or r.get("Email", "") or "").strip()
            roster_contact = str(
                r.get("CONTACT_NUMBER", "")
                or r.get("CONTACT", "")
                or r.get("MOBILE", "")
                or r.get("PHONE", "")
                or ""
            ).strip()
            if roster_email:
                st.session_state["email__pending"] = roster_email
            if roster_contact:
                st.session_state["contact_number__pending"] = roster_contact


            # Team fields:
            # The form now uses Team Name as the selected widget and shows Team Code automatically.
            # Therefore we must set BOTH the code-related state and the team_name_selected widget key.
            tcode_pick = _match_option_case_insensitive(tcode_raw, TEAM_CODES)
            resolved_team_code = tcode_pick or tcode_raw
            resolved_team_name = tname or (get_team_name(resolved_team_code) if resolved_team_code else "")

            if tcode_pick:
                st.session_state["team_code__pending"] = tcode_pick
                st.session_state["team_code_override__pending"] = ""
            else:
                st.session_state["team_code__pending"] = tcode_raw
                st.session_state["team_code_override__pending"] = tcode_raw

            if resolved_team_name:
                st.session_state["team_name_selected__pending"] = resolved_team_name
                st.session_state["team_name_override__pending"] = resolved_team_name
            else:
                st.session_state["team_name_selected__pending"] = ""
                st.session_state["team_name_override__pending"] = ""

            st.session_state["athlete_roster_match__pending"] = "(keep typed)"
            st.rerun()


# Combined name (display)
typed_full_name = " ".join([p for p in [first_name, other_name, last_name] if (p or "").strip()]).strip()
db_name_override = (st.session_state.get("db_name_override", "") or "").strip()

# Full Name (auto) — editable
full_name_display = (st.session_state.get("full_name", "") or "").strip()
if (not full_name_display) and typed_full_name:
    # Pre-fill from typed First/Other/Last (user can edit)
    st.session_state["full_name"] = typed_full_name
    full_name_display = typed_full_name
st.text_input("Full Name (auto)", key="full_name")

# Live validation: name presence
selected_from_roster = bool((st.session_state.get("unique_id_override", "") or "").strip() or (db_name_override or "").strip())
first_last_ok = selected_from_roster or (bool((first_name or "").strip()) and bool((last_name or "").strip()))
name_ok = passport_ok and first_last_ok
if not first_last_ok:
    st.warning("First Name and Last Name are required unless you selected the athlete from the roster.")


c4, c5, c6 = st.columns(3)
with c4:
    # Birth Date input (allow rendering even if a preloaded value is later than today)
    _birth_cur = st.session_state.get("birth_date")
    _birth_max = dt.date.today()
    if isinstance(_birth_cur, dt.date) and _birth_cur > _birth_max:
        _birth_max = _birth_cur
    birth_date = st.date_input(
        "Birth Date",
        value=_birth_cur if isinstance(_birth_cur, dt.date) else None,
        min_value=dt.date(1900, 1, 1),
        max_value=_birth_max,
        key="birth_date",
        format="DD-MM-YYYY",
    )
    # Live validation: birth date
    birth_ok = (st.session_state.get("birth_date") is not None) and (st.session_state.get("birth_date") <= dt.date.today())
    if st.session_state.get("birth_date") and st.session_state.get("birth_date") > dt.date.today():
        st.warning("Birth Date cannot be in the future.")
    elif not birth_ok:
        st.warning("Birth Date is required.")

with c6:
    nationality_options = [""] + (COUNTRIES or [])
    _nat_extra = (st.session_state.get("nationality_override", "") or "").strip()
    if _nat_extra and _nat_extra not in nationality_options:
        nationality_options = ["", _nat_extra] + [x for x in nationality_options if x != ""]
    nationality = st.selectbox("Nationality", nationality_options, index=0, key="nationality")
    is_singapore = (str(nationality or '').strip().upper() in ('SGP','SIN','SG','SINGAPORE'))
    # Singapore PR status (separate from nationality code)
    singapore_pr = st.checkbox('Singapore PR?', key='singapore_pr')



# Unique ID (display) — placed under Birth Date / IC row
unique_id_override = (st.session_state.get("unique_id_override", "") or "").strip()
_ic_for_uid = normalize_ic_last4(st.session_state.get("ic_last4", "") or "")
unique_id = unique_id_override or (compute_unique_id(first_name, _ic_for_uid, birth_date) if birth_date else "")
uid_from_roster = unique_id_override
if uid_from_roster:
    st.text_input("Unique ID (from roster)", value=uid_from_roster, disabled=True)
else:
    st.text_input("Unique ID (auto)", value=(unique_id or ""), disabled=True)


with c5:
    ic_last4 = st.text_input("IC Number (last 4)", key="ic_last4")
    # Live validation: IC last-4 (3 digits + 1 letter)
    ic_last4_norm = normalize_ic_last4(ic_last4)  # ALWAYS define
    # IC last-4 is required if Singapore PR is ticked, or if Singapore athlete has no UNIQUE_ID
    unique_id_present_for_ic = bool((st.session_state.get("unique_id_override", "") or "").strip() or (st.session_state.get("unique_id", "") or "").strip())
    ic_required = bool(singapore_pr) or (bool(is_singapore) and (not unique_id_present_for_ic))
    ic_ok = True
    if ic_required and not ic_last4_norm:
        ic_ok = False
        st.warning("IC format: 3 digits + 1 letter (e.g., 123A) — required when Singapore PR is ticked, or when a Singapore athlete has no UNIQUE_ID.")
    elif (not ic_last4_norm):
        # Not required and not provided
        ic_ok = True
    elif len(ic_last4_norm) < 4:
        ic_ok = False
        st.warning("IC last 4 is incomplete (e.g., 123A).")
    else:
        ic_ok = is_valid_ic_last4(ic_last4_norm)
        if not ic_ok:
            st.error("IC last 4 must be 3 digits followed by 1 letter (e.g., 123A).")

c7, c8 = st.columns(2)
contact_number = c7.text_input("Contact Number", key="contact_number")

# Live validation: contact number
contact_ok = bool((contact_number or '').strip())
if not contact_ok:
    st.warning("Contact Number is required.")

email = c8.text_input("Email", key="email")

# Live validation: email
email_norm = normalize_email(email)
email_ok = True
if email_norm:
    email_ok = is_valid_email(email_norm)

email_present = bool(email_norm)
if not email_present:
    st.warning("Email is required.")
    if not email_ok:
        st.error("Please enter a valid email address (e.g., name@example.com).")


c9, c10 = st.columns(2)
# Per-entry team selection; user selects Team Name, Team Code is auto-shown
team_name_by_code = {code: (get_team_name(code) or code) for code in TEAM_CODES}
team_code_by_name = {}
team_name_options = []
for _code, _name in team_name_by_code.items():
    if _name and _name not in team_code_by_name:
        team_code_by_name[_name] = _code
        team_name_options.append(_name)

_tc_extra = (st.session_state.get("team_code_override", "") or st.session_state.get("team_code", "") or "").strip()
_tn_extra = (st.session_state.get("team_name_override", "") or "").strip()
if _tn_extra and _tn_extra not in team_code_by_name:
    team_name_options = [_tn_extra] + team_name_options
    team_code_by_name[_tn_extra] = _tc_extra or ""
elif _tc_extra and _tc_extra in TEAM_CODES:
    _tn_from_code = get_team_name(_tc_extra) or _tc_extra
    if _tn_from_code and _tn_from_code not in team_code_by_name:
        team_name_options = [_tn_from_code] + team_name_options
        team_code_by_name[_tn_from_code] = _tc_extra

# Apply default/sidebar/roster value before the widget is created
_default_team_name_selected = ""
if _tn_extra:
    _default_team_name_selected = _tn_extra
elif _tc_extra:
    _default_team_name_selected = get_team_name(_tc_extra) or _tc_extra
elif default_team_code and default_team_code in TEAM_CODES:
    _default_team_name_selected = get_team_name(default_team_code) or default_team_code

if _default_team_name_selected and _default_team_name_selected in team_name_options:
    if st.session_state.get("team_name_selected") not in team_name_options:
        st.session_state["team_name_selected"] = _default_team_name_selected

team_name_row = c9.selectbox("Team Name", team_name_options, key="team_name_selected")
team_code = team_code_by_name.get(team_name_row, "")
st.session_state["team_code"] = team_code
c10.text_input("Team Code (auto)", team_code, disabled=True)

c11, c12 = st.columns(2)
event_division = c11.selectbox(
    "Event Division (1–8)",
    options=list(DIVISIONS.keys()),
    format_func=lambda k: f"{k} - {DIVISIONS[k]}",
    key="event_division",
)

event_opts_raw = allowed_events(gender_to_code(gender), int(event_division))
# De-duplicate event names (some divisions have duplicates)
event_opts = []
_seen_event_names = set()
for _n, _c in (event_opts_raw or []):
    if _n not in _seen_event_names:
        event_opts.append((_n, _c))
        _seen_event_names.add(_n)
event_opts = sorted(event_opts, key=lambda _x: _event_sort_key(_x[0]))
event_names = [n for n, _ in event_opts]

# Keep event selection consistent when options change (multi-select)
prev_selected = st.session_state.get("events_selected", [])
if not isinstance(prev_selected, list):
    prev_selected = []
prev_selected_valid = [e for e in prev_selected if e in event_names]
if prev_selected_valid != prev_selected:
    st.session_state["events_selected"] = prev_selected_valid

selected_events = c12.multiselect(
    "Select event(s)",
    options=event_names,
    default=prev_selected_valid,
    key="events_selected",
)

# Live validation: must have at least one event selected
event_ok = bool(event_opts) and len(selected_events) > 0
if not event_ok:
    st.warning("Please select at least one event for the selected division.")


season_best = st.text_input("Season Best", key="season_best")
season_best_ok = bool((season_best or "").strip())
if not season_best_ok:
    st.warning("Season Best is required.")
emergency_contact_name = st.text_input("Emergency Contact Name", key="emergency_contact_name")
emergency_contact_number = st.text_input("Emergency Contact Number", key="emergency_contact_number")
coach_full_name = st.text_input("Coach Full Name", key="coach_full_name")
parq = st.selectbox("PAR-Q completed?", ["Y", "N"], key="parq")

ic_last4_norm = normalize_ic_last4(ic_last4)
email_norm = normalize_email(email)

waiver_ok = st.checkbox("I acknowledge the waiver (as per the original form).", value=False, key="waiver_ok")

# Gate Add entry button (live checks)
ready_to_add = bool(waiver_ok) and bool(email_present) and bool(email_ok) and bool(ic_ok) and bool(birth_ok) and bool(contact_ok) and bool(name_ok) and bool(gender_ok) and bool(event_ok) and bool(season_best_ok)



# Add entry button
if st.button("Add entry", type="primary", disabled=not ready_to_add):
    missing = []
    _uid_present = bool((unique_id or "").strip())
    _is_sgp_local = (str(nationality or "").strip().upper() in ("SGP","SIN","SG","SINGAPORE"))
    _pr_local = bool(st.session_state.get("singapore_pr", False))
    ic_required = bool(_pr_local) or (bool(_is_sgp_local) and (not _uid_present))
    missing_checks = [
        ("Name as per NRIC/Passport", (st.session_state.get("name_passport","") or "").strip()),
        ("Birth Date", birth_date),
        ("Email", email),
        ("Contact Number", contact_number),
        ("Season Best", season_best),
    ]
        # IC required only if Singapore PR is ticked OR Singapore athlete has no UNIQUE_ID
    if ic_required:
        missing_checks.insert(1, ("IC last 4", ic_last4))
    for k, v in missing_checks:

        if not v:
            missing.append(k)

    if not waiver_ok:
        st.error("Please tick the waiver acknowledgement.")
    elif missing:
        st.error("Missing: " + ", ".join(missing))
    elif not gender_ok:
        st.error("Please select Gender (Male or Female).")
    elif not ((st.session_state.get("name_passport","") or "").strip()):
        st.error("Name as per NRIC/Passport is required.")
    elif (not ((st.session_state.get("unique_id_override","") or "").strip() or (db_name_override or "").strip())) and (not (((first_name or "").strip()) and ((last_name or "").strip()))):
        st.error("First Name and Last Name are required unless you selected the athlete from the roster.")
    elif not is_valid_email(email_norm):
        st.error("Please enter a valid email address (e.g., name@example.com).")
    elif (str(nationality or '').strip().upper() in ('SGP','SIN','SG','SINGAPORE')) and (not _uid_present) and (not is_valid_ic_last4(ic_last4_norm)):
        st.error("IC last 4 must be 3 digits followed by 1 letter (e.g., 123A).")
    elif _uid_present and ic_last4_norm and (not is_valid_ic_last4(ic_last4_norm)):
        st.error("IC last 4 must be 3 digits followed by 1 letter (e.g., 123A).")
    elif not event_opts or not event_names or not selected_events:
        st.error("Please select at least one event for that Gender + Division combination.")
    elif not (season_best or "").strip():
        st.error("Season Best is required.")
    else:
        # Add one row per selected event
        added_events = []
        for _ev in selected_events:
            _code = dict(event_opts).get(_ev, "")
            st.session_state.entries.append({
            "name": (db_name_override or typed_full_name),
            "full_name": (st.session_state.get("full_name", "") or db_name_override or typed_full_name),
            "name_passport": (st.session_state.get("name_passport", "") or "").strip(),
            "last_name": (last_name or "").strip(),
            "first_name": (first_name or "").strip(),
            "other_name": (other_name or "").strip(),
            "gender": gender,
            "birth_date": birth_date,
            "ic_last4": ic_last4_norm,
            "unique_id": unique_id,
            "nationality": nationality,
            "singapore_pr": singapore_pr,
            "contact_number": (contact_number or "").strip(),
            "email": email_norm,
            "team_code": team_code,
            "team_name": team_name_row,
            "charge_code": charge_code,
            "po_to_be_sent": po_to_be_sent,
            "event_division": int(event_division),
            "season_best": (season_best or "").strip(),
            "emergency_contact_name": (emergency_contact_name or "").strip(),
            "emergency_contact_number": (emergency_contact_number or "").strip(),
            "coach_full_name": (coach_full_name or "").strip(),
            "parq": parq,
            "event": _ev,
            "event_code": _code,
            })
            added_events.append(_ev)
        st.success(f"Added {len(added_events)} entry(ies).")
        # Send confirmation email (SMTP) — do not block saving if email fails
        try:
            if email_norm and is_valid_email(email_norm):
                _subj = "Entry confirmation"
                _full = (st.session_state.get("full_name", "") or "").strip() or (db_name_override or typed_full_name)
                _uid_disp = (st.session_state.get("unique_id_override", "") or "").strip() or unique_id
                _body = (
                    "Dear Participant,\n\n"
                    "Your entry has been successfully received.\n\n"
                    f"Full Name: {_full}\n"
                    f"Event(s): {', '.join(added_events)}\n"
                    f"Team: {team_name_row}\n"
                    f"Unique ID: {_uid_disp}\n\n"
                    "Thank you.\n\n"
                    "SAA\n"
                )
                st.session_state["email_last_attempt"] = {
                    "ts": dt.datetime.utcnow().isoformat() + "Z",
                    "to": email_norm,
                    "subject": _subj,
                    "events": list(added_events) if isinstance(added_events, list) else str(added_events),
                    "status": "attempting",
                }
                send_confirmation_email_smtp(email_norm, _subj, _body)
                st.session_state["email_last_attempt"]["status"] = "sent"
                st.toast("Confirmation email sent.")
        except Exception as e:
            st.session_state["email_last_attempt"] = {
                "ts": dt.datetime.utcnow().isoformat() + "Z",
                "to": email_norm,
                "status": "failed",
                "error_type": type(e).__name__,
                "error": repr(e),
                "traceback": tb.format_exc(),
            }
            st.warning(f"Entry added, but email failed: {type(e).__name__}: {repr(e)}")
            with st.expander("Email error traceback"):
                st.code(st.session_state["email_last_attempt"]["traceback"])


        # Sync "Current entries" to output Google Sheet (optional)
        if st.session_state.get("sync_enabled") and (st.session_state.get("output_sheet_url") or "").strip():
            try:
                sync_entries_to_sheet(
                    st.session_state.entries,
                    sheet_url_or_id=st.session_state.get("output_sheet_url", ""),
                    worksheet=((st.session_state.get("output_worksheet") or "").strip() or None),
                )
                st.toast("Synced to output Google Sheet.")
            except Exception as e:
                st.warning(f"Output sheet sync failed: {type(e).__name__}: {repr(e)}")


        # Auto-clear form fields for the next entry (use __pending to avoid Streamlit widget-state mutation errors)
        for _k, _v in {
            "last_name": "",
            "first_name": "",
            "other_name": "",
            "gender": "",
            "birth_date": None,
            "ic_last4": "",
            "contact_number": "",
            "email": "",
            "season_best": "",
            "emergency_contact_name": "",
            "emergency_contact_number": "",
            "coach_full_name": "",
            "waiver_ok": False,
            "db_name_override": "",
            "athlete_roster_match": "(keep typed)",
            "events_selected": [],
            "singapore_pr": False,
        }.items():
            st.session_state[f"{_k}__pending"] = _v
        st.rerun()



# -------- Public entry-only mode --------
# Existing/current entries, download buttons, and edit/delete controls are intentionally hidden
# in this app variant. Entries are still saved to session state and synced to the output
# Google Sheet when "Add entry" succeeds above.
st.caption("Entry-only mode: existing entries and edit controls are hidden.")
